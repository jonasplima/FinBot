"""Tests for ExportService."""

import base64
import io
from datetime import date
from unittest.mock import AsyncMock

import pytest
from openpyxl import load_workbook

from app.services.export import ExportService
from app.utils.validators import sanitize_for_spreadsheet


class TestExportServiceExportMonth:
    """Tests for ExportService.export_month method."""

    @pytest.fixture
    def service(self):
        return ExportService()

    async def test_export_month_no_expenses(self, service, seeded_session, test_phone):
        """Test export when there are no expenses."""
        result = await service.export_month(seeded_session, test_phone)

        assert result["success"] is False
        assert "nao tem gastos" in result["message"].lower()

    async def test_export_month_with_expenses(
        self, service, seeded_session, test_phone, expense_in_db
    ):
        """Test successful export with expenses."""
        result = await service.export_month(seeded_session, test_phone)

        assert result["success"] is True
        assert "file_base64" in result
        assert "filename" in result
        assert result["filename"].endswith(".xlsx")

    async def test_export_generates_valid_xlsx(
        self, service, seeded_session, test_phone, expense_in_db
    ):
        """Test that the exported file is a valid XLSX."""
        result = await service.export_month(seeded_session, test_phone)

        assert result["success"] is True

        # Decode and load the XLSX
        file_bytes = base64.b64decode(result["file_base64"])
        wb = load_workbook(io.BytesIO(file_bytes))

        # Check that it has content
        ws = wb.active
        assert ws is not None

        # Check headers in first row
        headers = [cell.value for cell in ws[1]]
        assert "Data" in headers
        assert "Descricao" in headers
        assert "Categoria" in headers
        assert "Valor" in headers

    async def test_export_specific_month(self, service, seeded_session, test_phone):
        """Test export for a specific month."""
        result = await service.export_month(seeded_session, test_phone, month=3, year=2024)

        # No data for that month
        assert result["success"] is False
        assert "Marco" in result["message"]
        assert "2024" in result["message"]

    async def test_export_filename_format(self, service, seeded_session, test_phone, expense_in_db):
        """Test that filename follows expected format."""
        result = await service.export_month(seeded_session, test_phone)

        today = date.today()
        assert result["success"] is True
        assert f"{today.year}" in result["filename"]
        assert result["filename"].startswith("gastos_")

    async def test_export_includes_totals(self, service, seeded_session, test_phone, expense_in_db):
        """Test that export includes total calculations."""
        result = await service.export_month(seeded_session, test_phone)

        assert result["success"] is True

        # Load and check for totals
        file_bytes = base64.b64decode(result["file_base64"])
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb.active

        # Find TOTAL row
        found_total = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "TOTAL":
                    found_total = True
                    break

        assert found_total

    async def test_export_month_pdf_with_expenses(
        self, service, seeded_session, test_phone, expense_in_db
    ):
        """Test successful PDF export with expenses."""
        result = await service.export_month_pdf(seeded_session, test_phone)

        assert result["success"] is True
        assert "file_base64" in result
        assert result["filename"].endswith(".pdf")
        assert result["mimetype"] == "application/pdf"

    async def test_export_month_pdf_generates_valid_pdf(
        self, service, seeded_session, test_phone, expense_in_db
    ):
        """Test that the exported file is a valid PDF."""
        result = await service.export_month_pdf(seeded_session, test_phone)

        assert result["success"] is True

        file_bytes = base64.b64decode(result["file_base64"])
        assert file_bytes.startswith(b"%PDF")
        assert len(file_bytes) > 1000

    async def test_export_month_pdf_no_expenses(self, service, seeded_session, test_phone):
        """Test PDF export when there are no expenses."""
        result = await service.export_month_pdf(seeded_session, test_phone)

        assert result["success"] is False
        assert "nao tem gastos" in result["message"].lower()


class TestExportServiceWithMultipleExpenses:
    """Tests for ExportService with multiple expenses."""

    @pytest.fixture
    def service(self):
        return ExportService()

    async def test_export_multiple_expenses(
        self, service, seeded_session, test_phone, sample_expense_data
    ):
        """Test export with multiple expenses."""
        from app.services.expense import ExpenseService

        expense_service = ExpenseService()

        # Create multiple expenses
        for i in range(3):
            data = sample_expense_data.copy()
            data["description"] = f"Expense {i + 1}"
            data["amount"] = 50.00 + (i * 10)
            await expense_service.create_expense(seeded_session, test_phone, data)

        result = await service.export_month(seeded_session, test_phone)

        assert result["success"] is True

        # Verify all expenses are in the file
        file_bytes = base64.b64decode(result["file_base64"])
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb.active

        # Count data rows (excluding header and summary)
        data_rows = 0
        for row in ws.iter_rows(min_row=2):
            if row[0].value and row[0].value != "TOTAL":
                data_rows += 1

        assert data_rows == 3

    async def test_export_calculates_correct_totals(
        self, service, seeded_session, test_phone, sample_expense_data
    ):
        """Test that export calculates correct totals."""
        from app.services.expense import ExpenseService

        expense_service = ExpenseService()

        # Create expenses with known amounts
        amounts = [50.00, 75.00, 100.00]
        for i, amount in enumerate(amounts):
            data = sample_expense_data.copy()
            data["description"] = f"Expense {i + 1}"
            data["amount"] = amount
            await expense_service.create_expense(seeded_session, test_phone, data)

        result = await service.export_month(seeded_session, test_phone)

        assert result["success"] is True

        # Verify totals in the file
        file_bytes = base64.b64decode(result["file_base64"])
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb.active

        # Find the total value
        expected_total = sum(amounts)
        found_correct_total = False

        for row in ws.iter_rows():
            for cell in row:
                if cell.value == expected_total:
                    found_correct_total = True
                    break

        assert found_correct_total


class TestSpreadsheetSanitization:
    """Tests for spreadsheet-safe export behavior."""

    @pytest.fixture
    def service(self):
        return ExportService()

    @pytest.mark.parametrize(
        ("raw_value", "expected"),
        [
            ("=SUM(A1:A2)", "'=SUM(A1:A2)"),
            ("+cmd", "'+cmd"),
            ("-10", "'-10"),
            ("@user", "'@user"),
            ("almoco", "almoco"),
        ],
    )
    def test_sanitize_for_spreadsheet(self, raw_value, expected):
        """Test spreadsheet sanitizer neutralizes formula-like prefixes."""
        assert sanitize_for_spreadsheet(raw_value) == expected

    async def test_export_sanitizes_formula_like_description(
        self, service, seeded_session, test_phone, sample_expense_data
    ):
        """Test XLSX export neutralizes formula-like text fields."""
        service.expense_service.get_expenses_for_export = AsyncMock(
            return_value=[
                {
                    "Data": "2026-04-06",
                    "Descricao": "=SUM(A1:A2)",
                    "Categoria": "@mercado",
                    "Forma de Pagamento": "+pix",
                    "Tipo": "Negativo",
                    "Parcela": "-1/3",
                    "Valor": 50.0,
                    "Compartilhada": "Nao",
                    "Percentual": 0,
                }
            ]
        )

        result = await service.export_month(seeded_session, test_phone)

        assert result["success"] is True

        file_bytes = base64.b64decode(result["file_base64"])
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb.active

        assert ws["B2"].value == "'=SUM(A1:A2)"
        assert ws["C2"].value == "'@mercado"
        assert ws["D2"].value == "'+pix"

    async def test_export_preserves_numeric_value_column(
        self, service, seeded_session, test_phone, sample_expense_data
    ):
        """Test XLSX export keeps monetary values as numbers."""
        from app.services.expense import ExpenseService

        expense_service = ExpenseService()
        data = sample_expense_data.copy()
        data["description"] = "-mercado"
        data["amount"] = 123.45
        await expense_service.create_expense(seeded_session, test_phone, data)

        result = await service.export_month(seeded_session, test_phone)

        assert result["success"] is True

        file_bytes = base64.b64decode(result["file_base64"])
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb.active

        assert ws["B2"].value == "'-mercado"
        assert ws["G2"].value == pytest.approx(123.45)
        assert isinstance(ws["G2"].value, float)

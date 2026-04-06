"""Export service for generating XLSX and PDF files."""

import base64
import io
import logging
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.chart import ChartService
from app.services.expense import MONTH_NAMES, ExpenseService

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting expenses to XLSX."""

    def __init__(self):
        self.expense_service = ExpenseService()
        self.chart_service = ChartService()

    def _resolve_period(
        self,
        month: int | None,
        year: int | None,
    ) -> tuple[int, int, str]:
        """Resolve export period, defaulting to current month/year."""
        today = date.today()
        resolved_month = month if month is not None else today.month
        resolved_year = year if year is not None else today.year
        month_name = MONTH_NAMES.get(resolved_month, str(resolved_month))
        return resolved_month, resolved_year, month_name

    def _format_currency(self, value: float | Decimal) -> str:
        """Format currency values in pt-BR style."""
        amount = float(value)
        return f"R$ {amount:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def _calculate_totals(self, expenses: list[dict]) -> tuple[float, float, float]:
        """Calculate negative, positive and balance totals from export rows."""
        total_negativo = sum(
            expense["Valor"] for expense in expenses if expense.get("Tipo") == "Negativo"
        )
        total_positivo = sum(
            expense["Valor"] for expense in expenses if expense.get("Tipo") == "Positivo"
        )
        saldo = total_positivo - total_negativo
        return total_negativo, total_positivo, saldo

    async def export_month(
        self,
        session: AsyncSession,
        phone: str,
        month: int | None = None,
        year: int | None = None,
    ) -> dict:
        """
        Export expenses for a month to XLSX.

        Returns:
            dict with success, file_base64, filename, month_name
        """
        month, year, month_name = self._resolve_period(month, year)

        # Get expenses
        expenses = await self.expense_service.get_expenses_for_export(session, phone, month, year)

        if not expenses:
            return {
                "success": False,
                "message": f"Voce nao tem gastos em {month_name} de {year}.",
            }

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{month_name} {year}"

        # Define headers
        headers = [
            "Data",
            "Descricao",
            "Categoria",
            "Forma de Pagamento",
            "Tipo",
            "Parcela",
            "Valor",
            "Compartilhada",
            "Percentual",
        ]

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data
        for row_idx, expense in enumerate(expenses, 2):
            for col_idx, header in enumerate(headers, 1):
                value = expense.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # Format numbers
                if header == "Valor":
                    cell.number_format = "R$ #,##0.00"
                elif header == "Percentual" and value:
                    cell.number_format = "0.00%"

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col)

            for row in range(1, len(expenses) + 2):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add summary row
        summary_row = len(expenses) + 3
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)

        # Calculate totals
        total_negativo, total_positivo, saldo = self._calculate_totals(expenses)

        ws.cell(row=summary_row, column=6, value="Gastos:").font = Font(bold=True)
        ws.cell(row=summary_row, column=7, value=total_negativo).number_format = "R$ #,##0.00"

        ws.cell(row=summary_row + 1, column=6, value="Entradas:").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=7, value=total_positivo).number_format = "R$ #,##0.00"

        ws.cell(row=summary_row + 2, column=6, value="Saldo:").font = Font(bold=True)
        saldo_cell = ws.cell(row=summary_row + 2, column=7, value=saldo)
        saldo_cell.number_format = "R$ #,##0.00"
        saldo_cell.font = Font(bold=True)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Encode to base64
        file_base64 = base64.b64encode(output.getvalue()).decode("utf-8")

        filename = f"gastos_{month_name.lower()}_{year}.xlsx"

        return {
            "success": True,
            "file_base64": file_base64,
            "filename": filename,
            "month_name": f"{month_name} de {year}",
        }

    async def export_month_pdf(
        self,
        session: AsyncSession,
        phone: str,
        month: int | None = None,
        year: int | None = None,
    ) -> dict:
        """
        Export expenses for a month to PDF.

        Returns:
            dict with success, file_base64, filename, month_name, mimetype
        """
        month, year, month_name = self._resolve_period(month, year)

        expenses = await self.expense_service.get_expenses_for_export(session, phone, month, year)

        if not expenses:
            return {
                "success": False,
                "message": f"Voce nao tem gastos em {month_name} de {year}.",
            }

        category_data = await self.expense_service.get_expenses_by_category(
            session, phone, month, year
        )
        total_negativo, total_positivo, saldo = self._calculate_totals(expenses)

        output = io.BytesIO()
        document = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
        )

        styles = getSampleStyleSheet()
        title_style = styles["Heading1"]
        subtitle_style = styles["Heading2"]
        body_style = styles["BodyText"]
        body_style.fontName = "Helvetica"
        body_style.leading = 14
        summary_style = ParagraphStyle(
            "Summary",
            parent=body_style,
            fontSize=11,
            leading=15,
            spaceAfter=6,
        )

        story = [
            Paragraph(f"Relatorio Financeiro - {month_name} de {year}", title_style),
            Spacer(1, 0.3 * cm),
            Paragraph("Resumo do periodo", subtitle_style),
            Paragraph(f"Gastos: {self._format_currency(total_negativo)}", summary_style),
            Paragraph(f"Entradas: {self._format_currency(total_positivo)}", summary_style),
            Paragraph(f"Saldo: {self._format_currency(saldo)}", summary_style),
            Spacer(1, 0.4 * cm),
        ]

        if category_data:
            chart_bytes = self.chart_service.generate_pie_chart(
                category_data,
                title=f"Gastos por Categoria - {month_name} de {year}",
            )
            chart_buffer = io.BytesIO(chart_bytes)
            chart_image = Image(chart_buffer, width=14 * cm, height=10 * cm)
            story.extend(
                [
                    Paragraph("Distribuicao por categoria", subtitle_style),
                    chart_image,
                    Spacer(1, 0.5 * cm),
                ]
            )

        story.append(Paragraph("Lancamentos", subtitle_style))

        table_data = [
            [
                "Data",
                "Descricao",
                "Categoria",
                "Pagamento",
                "Tipo",
                "Valor",
            ]
        ]
        for expense in expenses:
            table_data.append(
                [
                    expense["Data"],
                    expense["Descricao"],
                    expense["Categoria"],
                    expense["Forma de Pagamento"],
                    expense["Tipo"],
                    self._format_currency(expense["Valor"]),
                ]
            )

        table = Table(
            table_data,
            repeatRows=1,
            colWidths=[2.1 * cm, 5.5 * cm, 3.3 * cm, 3.1 * cm, 2.2 * cm, 2.8 * cm],
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#F7F9FC")],
                    ),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (5, 1), (5, -1), "RIGHT"),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(table)

        document.build(story)
        output.seek(0)

        file_base64 = base64.b64encode(output.getvalue()).decode("utf-8")
        filename = f"gastos_{month_name.lower()}_{year}.pdf"

        return {
            "success": True,
            "file_base64": file_base64,
            "filename": filename,
            "month_name": f"{month_name} de {year}",
            "mimetype": "application/pdf",
        }
